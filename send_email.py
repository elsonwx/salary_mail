#!/usr/bin/env python
# -*-coding:utf-8-*-

import os
import sys
import codecs
import time
import datetime
import smtplib
import configparser
from email.mime.text import MIMEText
from email.header import Header
from openpyxl import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
log_path = current_dir + os.sep + 'log.txt'


def loginfo(msg):
    with codecs.open(log_path, 'a', 'utf-8') as f:
        f.write(time.strftime("%Y-%m-%d %X") + "-" + msg + os.linesep)


def send_mail(to_addr, subject, html_template, user_mail, user_passwd, smtp_server, smtp_port, enable_ssl):
    try:
        message = MIMEText(html_template, 'html', 'utf-8')
        message['From'] = Header(user_mail)
        message['To'] = Header(to_addr)
        message['Subject'] = Header(subject, 'utf-8')
        mail_obj = None
        if enable_ssl:
            mail_obj = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            mail_obj = smtplib.SMTP(smtp_server, smtp_port)
        mail_obj.login(user_mail, user_passwd)
        mail_obj.sendmail(user_mail, to_addr, message.as_string())
        mail_obj.quit()
        return True
    except Exception as e:
        loginfo('send mail to ' + str(to_addr) + ' failed,exception: ' + str(e))
        return False


def read_data(excel_file):
    excel_data = []
    # occupy lines for each item(table header,staff salary,staff salary....)
    item_lines_arr = []
    wb = load_workbook(filename=excel_file, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.rows:
        row_cells = []
        for index, cell in enumerate(row):
            cell_merge = get_cell_merge(cell.row, cell.column, ws.merged_cells)
            if index == 0:
                if cell_merge["type"] == 'rowspan':
                    item_lines_arr.append(cell_merge["rowspan"])
                elif cell_merge["type"] == 'normal':
                    item_lines_arr.append(1)
            row_cells.append({
                "value": cell.value,
                "coordinate": cell.coordinate,
                "col": cell.column,
                "row": cell.row,
                "merge": cell_merge
            })
        excel_data.append(row_cells)
    return excel_data, item_lines_arr


def read_attach():
    attach_path = current_dir + os.sep + 'attach.txt'
    if os.path.exists(attach_path):
        with open(attach_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    else:
        return ''


def get_cell_merge(row, col, merged_cells):
    for item in merged_cells.ranges:
        # on the same column
        if item.min_col == item.max_col == col:
            # rowspan
            if item.min_row == row:
                return {"type": "rowspan", "rowspan": item.max_row - item.min_row + 1}
            elif item.min_row < row <= item.max_row:
                return {"type": "none"}
        # on the same row
        elif item.max_row == item.min_row == row:
            # colspan
            if item.min_col == col:
                return {"type": "colspan", "colspan": item.max_col - item.min_col + 1}
            elif item.min_col < col <= item.max_col:
                return {"type": "none"}
        elif item.min_row == row and item.min_col == col:
            return {"type": "mix", "rowspan": item.max_row - item.min_row + 1,
                    "colspan": item.max_col - item.min_col + 1}
        elif item.min_row <= row <= item.max_row and item.min_col <= col <= item.max_col:
            return {"type": "none"}
    return {"type": "normal"}


def fill_table(row_datas, style):
    grid = 'td' if style == 'td' else 'th'
    holder_str = ''
    for row_cells in row_datas:
        holder_str += '<tr>'
        for cell in row_cells[1:]:
            try:
                val = '' if cell["value"] is None else cell["value"]
            except Exception as e:
                print(e)
            if cell["merge"]["type"] == 'rowspan':
                holder_str += '<%s style="padding-left:20px;padding-right:20px;" rowspan="%s">%s</%s>'\
                              % (grid, cell["merge"]["rowspan"], val, grid)
            if cell["merge"]["type"] == 'colspan':
                holder_str += '<%s style="text-align:center;" colspan="%s">%s</%s>' \
                              % (grid, cell["merge"]["colspan"], val, grid)
            if cell["merge"]["type"] == 'mix':
                holder_str += '<%s style="text-align:center;" rowspan="%s" colspan="%s">%s</%s>'\
                              % (grid, cell["merge"]["rowspan"], cell["merge"]["colspan"], val, grid)
            if cell["merge"]["type"] == 'none':
                pass
            if cell["merge"]["type"] == 'normal':
                holder_str += '<%s style="padding-left:20px;padding-right:20px;">%s</%s>'\
                              % (grid, val, grid)
        holder_str += '</tr>'
    return holder_str


def main():
    cf = configparser.ConfigParser()
    cf.read(current_dir + os.sep + 'config.ini')
    user = cf.get('user', 'email')
    pwd = cf.get('user', 'password')
    server = cf.get('user', 'smtp_server')
    port = cf.getint('user', 'smtp_port')
    enable_ssl = cf.getboolean('user', 'enable_ssl')

    today_day = datetime.datetime.now().day
    today_month = datetime.datetime.now().month
    print('The Company paid wages before the 10th')
    print('Today is ' + time.strftime("%B %d"))
    mail_subject = "%s月份工资条，请查收"
    # Pay money before the 10th of each month
    if today_day > 10:
        mail_subject = mail_subject % today_month
    else:
        today_month = today_month - 1
        if today_month == 0:
            today_month = 12
        mail_subject = mail_subject % today_month
    english_month = datetime.date(1900, today_month, 1).strftime('%B')
    print('The mail subject will be show as "' + english_month + ' salley bill"')
    print("\n")

    html_template = '<pre>' + read_attach() + '</pre>'
    html_template += '<br/><br/><table border="1" style="border-collapse:collapse">'
    html_template += '<thead>'
    html_template += '<<header_placeholder>>'
    html_template += '</thead>'
    html_template += '<tbody>'
    html_template += '<<salary_placeholder>>'
    html_template += '</tbody>'
    html_template += '</table>'

    excel_data, item_lines_arr = read_data(current_dir + os.sep + 'data.xlsx')
    header_datas = excel_data[0:item_lines_arr[0]]
    holder_str = fill_table(header_datas, 'th')
    html_template = html_template.replace('<<header_placeholder>>', holder_str)

    has_failture = False
    staff_index = item_lines_arr[0]
    for staff_lines in item_lines_arr[1:]:
        staff_email = excel_data[staff_index][0]["value"]
        staff_datas = excel_data[staff_index:staff_index + staff_lines]
        holder_str = fill_table(staff_datas, 'td')
        html_content = html_template.replace('<<salary_placeholder>>', holder_str)
        if staff_email is not None:
            staff_email = staff_email.replace("\n", "").replace("\r", "").replace(" ", "")
            send_result = send_mail(staff_email, mail_subject, html_content, user, pwd, server, port, enable_ssl)
            if not send_result:
                has_failture = True
                print('mail to:' + str(staff_email) + ' failed!!!,please send this email manually.')
            else:
                print('mail to:' + str(staff_email) + ' Successfully')
                time.sleep(1)
        staff_index += staff_lines
    print("\n")
    if has_failture:
        print("There are some mails failed to be send, please check theme in the log.txt")
        print("\n")
        input('Please input any key to quit...')
    else:
        print("Program has run successfully,all the mails have been sent successfully.")
        print('The program will exit in 3 seconds...')
        time.sleep(3)
    sys.exit(0)


main()
